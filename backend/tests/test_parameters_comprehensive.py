"""
Comprehensive test suite for Parameters API
Tests parameter CRUD operations, Excel upload, and bulk operations
"""

import io

from fastapi import status
from openpyxl import Workbook


class TestParameterCRUD:
    """Test Parameter CRUD operations"""
    def test_create_parameter(self, client, auth_headers):
        """Test creating a new parameter"""
        response = client.post(
            "/api/parameters/",
            json={
                "name": "ALT_MSL",
                "description": "Altitude Mean Sea Level",
                "unit": "ft",
                "system": "Navigation",
                "category": "Position",
                "min_value": 0.0,
                "max_value": 50000.0
            },
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_201_CREATED
        data = response.json()
        assert data["name"] == "ALT_MSL"
        assert data["unit"] == "ft"
        assert data["system"] == "Navigation"
        assert "id" in data

    def test_create_duplicate_parameter(self, client, auth_headers):
        """Test creating a parameter with duplicate name"""
        # Create first parameter
        client.post(
            "/api/parameters/",
            json={
                "name": "DUPLICATE_PARAM",
                "description": "Test Parameter",
                "unit": "deg"
            },
            headers=auth_headers
        )

        # Try to create duplicate
        response = client.post(
            "/api/parameters/",
            json={
                "name": "DUPLICATE_PARAM",
                "description": "Another Test Parameter",
                "unit": "deg"
            },
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "already exists" in response.json()["detail"]

    def test_list_parameters(self, client, auth_headers):
        """Test listing all parameters"""
        # Create multiple parameters
        for i in range(5):
            client.post(
                "/api/parameters/",
                json={
                    "name": f"TEST_PARAM_{i}",
                    "description": f"Test Parameter {i}",
                    "unit": "unit"
                },
                headers=auth_headers
            )

        response = client.get("/api/parameters/", headers=auth_headers)

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert len(data) >= 5

    def test_list_parameters_pagination(self, client, auth_headers):
        """Test parameter listing with pagination"""
        # Create 10 parameters
        for i in range(10):
            client.post(
                "/api/parameters/",
                json={
                    "name": f"PAGINATION_PARAM_{i}",
                    "description": f"Pagination Test {i}",
                    "unit": "unit"
                },
                headers=auth_headers
            )

        # Test with limit
        response = client.get("/api/parameters/?limit=5",
                              headers=auth_headers)
        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert len(data) <= 5

        # Test with skip
        response = client.get("/api/parameters/?skip=5&limit=3",
                              headers=auth_headers)
        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert len(data) <= 3

    def test_get_parameter_by_id(self, client, auth_headers):
        """Test retrieving a specific parameter by ID"""
        # Create a parameter
        create_response = client.post(
            "/api/parameters/",
            json={
                "name": "SPECIFIC_PARAM",
                "description": "Specific Test Parameter",
                "unit": "deg",
                "system": "Flight Control"
            },
            headers=auth_headers
        )
        param_id = create_response.json()["id"]

        # Retrieve it
        response = client.get(f"/api/parameters/{param_id}",
                              headers=auth_headers)

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert data["id"] == param_id
        assert data["name"] == "SPECIFIC_PARAM"
        assert data["system"] == "Flight Control"

    def test_get_nonexistent_parameter(self, client, auth_headers):
        """Test retrieving a parameter that doesn't exist"""
        response = client.get("/api/parameters/99999", headers=auth_headers)
        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_update_parameter(self, client, auth_headers):
        """Test updating a parameter"""
        # Create a parameter
        create_response = client.post(
            "/api/parameters/",
            json={
                "name": "UPDATE_PARAM",
                "description": "Original Description",
                "unit": "deg"
            },
            headers=auth_headers
        )
        param_id = create_response.json()["id"]

        # Update it
        response = client.put(
            f"/api/parameters/{param_id}",
            json={
                "name": "UPDATE_PARAM",
                "description": "Updated Description",
                "unit": "rad",
                "system": "Updated System",
                "min_value": 0.0,
                "max_value": 360.0
            },
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert data["description"] == "Updated Description"
        assert data["unit"] == "rad"
        assert data["system"] == "Updated System"
        assert data["min_value"] == 0.0

    def test_delete_parameter(self, client, auth_headers):
        """Test deleting a parameter"""
        # Create a parameter
        create_response = client.post(
            "/api/parameters/",
            json={
                "name": "DELETE_PARAM",
                "description": "To Be Deleted",
                "unit": "unit"
            },
            headers=auth_headers
        )
        param_id = create_response.json()["id"]

        # Delete it
        response = client.delete(f"/api/parameters/{param_id}",
                                 headers=auth_headers)
        assert response.status_code == status.HTTP_204_NO_CONTENT

        # Verify it's gone
        get_response = client.get(f"/api/parameters/{param_id}",
                                  headers=auth_headers)
        assert get_response.status_code == status.HTTP_404_NOT_FOUND

    def test_search_parameters_by_system(self, client, auth_headers):
        """Test searching parameters by system"""
        # Create parameters with different systems
        systems = ["Navigation", "Flight Control", "Engine"]
        for _, system in enumerate(systems):
            for j in range(3):
                client.post(
                    "/api/parameters/",
                    json={
                        "name": f"{system.upper()}_{j}",
                        "description": f"{system} Parameter {j}",
                        "unit": "unit",
                        "system": system
                    },
                    headers=auth_headers
                )

        # Search for Navigation parameters
        response = client.get("/api/parameters/?system=Navigation",
                              headers=auth_headers)

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert all(p["system"] == "Navigation" for p in data)

    def test_search_parameters_by_category(self, client,
                                           auth_headers):
        """Test searching parameters by category"""
        # Create parameters with different categories
        categories = ["Position", "Attitude", "Performance"]
        for category in categories:
            for j in range(2):
                client.post(
                    "/api/parameters/",
                    json={
                        "name": f"{category.upper()}_{j}",
                        "description": f"{category} Parameter {j}",
                        "unit": "unit",
                        "category": category
                    },
                    headers=auth_headers
                )

        # Search for Position parameters
        response = client.get("/api/parameters/?category=Position",
                              headers=auth_headers)

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert all(p["category"] == "Position" for p in data)


class TestExcelUpload:
    """Test Excel file upload functionality"""

    def test_excel_upload_simple(self, client, auth_headers):
        """Test uploading a simple Excel file"""
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Parameters"

        # Add headers
        ws.append(["Name", "Description", "Unit", "System", "Category",
                   "Min Value", "Max Value"])

        # Add data
        ws.append(["ALT_MSL", "Altitude MSL", "ft", "Navigation",
                   "Position", 0, 50000])
        ws.append(["IAS", "Indicated Airspeed", "kt", "Navigation",
                   "Performance", 0, 500])
        ws.append(["PITCH", "Pitch Angle", "deg", "Flight Control",
                   "Attitude", -90, 90])

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Upload
        mime_type = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
        files = {"file": ("parameters.xlsx", excel_file, mime_type)}
        response = client.post(
            "/api/parameters/upload-excel",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert data["parameters_created"] >= 3
        assert "message" in data

    def test_excel_upload_with_duplicates(self, client, auth_headers):
        """Test Excel upload handling duplicate parameters"""
        # Create a parameter first
        client.post(
            "/api/parameters/",
            json={
                "name": "EXISTING_PARAM",
                "description": "Existing Parameter",
                "unit": "unit"
            },
            headers=auth_headers
        )

        # Create Excel with duplicate
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Description", "Unit", "System",
                   "Category", "Min Value", "Max Value"])
        ws.append(["EXISTING_PARAM", "Duplicate Parameter",
                   "unit", "System", "Category", 0, 100])
        ws.append(["NEW_PARAM", "New Parameter",
                   "unit", "System", "Category", 0, 100])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        mime_type = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
        files = {"file": ("parameters.xlsx", excel_file, mime_type)}
        response = client.post(
            "/api/parameters/upload-excel",
            files=files,
            headers=auth_headers
        )

        # Should handle duplicates gracefully
        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert "parameters_created" in data or "parameters_updated" in data

    def test_excel_upload_invalid_file_type(self, client, auth_headers):
        """Test uploading a non-Excel file"""
        # Create a text file
        text_file = io.BytesIO(b"Not an Excel file")

        files = {"file": ("test.txt", text_file,
                          "text/plain")}
        response = client.post(
            "/api/parameters/upload-excel",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Excel" in response.json()["detail"]

    def test_excel_upload_empty_file(self, client, auth_headers):
        """Test uploading an empty Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Description", "Unit"])  # Only headers, no data

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        mime_type = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
        files = {"file": ("empty.xlsx", excel_file, mime_type)}
        response = client.post(
            "/api/parameters/upload-excel",
            files=files,
            headers=auth_headers
        )

        # Should handle empty file gracefully
        assert response.status_code in [status.HTTP_200_OK,
                                        status.HTTP_400_BAD_REQUEST]

    def test_excel_upload_missing_required_columns(self, client, auth_headers):
        """Test Excel upload with missing required columns"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Description"])  # Missing Unit column
        ws.append(["TEST_PARAM", "Test Parameter"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        mime_type = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
        files = {"file": ("invalid.xlsx", excel_file, mime_type)}
        response = client.post(
            "/api/parameters/upload-excel",
            files=files,
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "required" in response.json()["detail"].lower()


class TestBulkOperations:
    """Test bulk parameter operations"""

    def test_bulk_create_parameters(self, client, auth_headers):
        """Test creating multiple parameters at once"""
        parameters = [
            {
                "name": f"BULK_PARAM_{i}",
                "description": f"Bulk Parameter {i}",
                "unit": "unit",
                "system": "Test System"
            }
            for i in range(10)
        ]

        response = client.post(
            "/api/parameters/bulk",
            json={"parameters": parameters},
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_201_CREATED
        data = response.json()
        assert data["created"] == 10

    def test_bulk_update_parameters(self, client, auth_headers):
        """Test updating multiple parameters at once"""
        # Create parameters first
        param_ids = []
        for i in range(5):
            response = client.post(
                "/api/parameters/",
                json={
                    "name": f"BULK_UPDATE_{i}",
                    "description": f"Original {i}",
                    "unit": "unit"
                },
                headers=auth_headers
            )
            param_ids.append(response.json()["id"])

        # Bulk update
        updates = [
            {
                "id": param_id,
                "description": f"Updated {i}",
                "system": "Updated System"
            }
            for i, param_id in enumerate(param_ids)
        ]

        response = client.put(
            "/api/parameters/bulk",
            json={"parameters": updates},
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert data["updated"] == 5

    def test_bulk_delete_parameters(self, client, auth_headers):
        """Test deleting multiple parameters at once"""
        # Create parameters first
        param_ids = []
        for i in range(5):
            response = client.post(
                "/api/parameters/",
                json={
                    "name": f"BULK_DELETE_{i}",
                    "description": f"To Delete {i}",
                    "unit": "unit"
                },
                headers=auth_headers
            )
            param_ids.append(response.json()["id"])

        # Bulk delete
        response = client.delete(
            "/api/parameters/bulk",
            json={"parameter_ids": param_ids},
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_204_NO_CONTENT

        # Verify all are deleted
        for param_id in param_ids:
            get_response = client.get(f"/api/parameters/{param_id}",
                                      headers=auth_headers)
            assert get_response.status_code == status.HTTP_404_NOT_FOUND


class TestParameterValidation:
    """Test parameter data validation"""

    def test_create_parameter_invalid_name(self, client, auth_headers):
        """Test creating parameter with invalid name"""
        response = client.post(
            "/api/parameters/",
            json={
                "name": "",  # Empty name
                "description": "Test",
                "unit": "unit"
            },
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_422_UNPROCESSABLE_ENTITY

    def test_create_parameter_missing_required_fields(self,
                                                      client, auth_headers):
        """Test creating parameter without required fields"""
        response = client.post(
            "/api/parameters/",
            json={
                "description": "Test"
                # Missing name and unit
            },
            headers=auth_headers
        )

        assert response.status_code == status.HTTP_422_UNPROCESSABLE_ENTITY

    def test_create_parameter_invalid_min_max(self, client, auth_headers):
        """Test creating parameter with min > max"""
        response = client.post(
            "/api/parameters/",
            json={
                "name": "INVALID_RANGE",
                "description": "Invalid Range",
                "unit": "unit",
                "min_value": 100.0,
                "max_value": 50.0  # Max less than min
            },
            headers=auth_headers
        )

        # Should either reject or swap values
        if response.status_code == status.HTTP_201_CREATED:
            data = response.json()
            # If accepted, values should be swapped
            assert data["min_value"] <= data["max_value"]
        else:
            assert response.status_code == status.HTTP_400_BAD_REQUEST


class TestAuthentication:
    """Test authentication requirements for parameter endpoints"""

    def test_create_parameter_without_auth(self, client):
        """Test that creating parameter requires authentication"""
        response = client.post(
            "/api/parameters/",
            json={
                "name": "UNAUTH_PARAM",
                "description": "Unauthorized",
                "unit": "unit"
            }
        )
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_list_parameters_without_auth(self, client):
        """Test that listing parameters requires authentication"""
        response = client.get("/api/parameters/")
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_upload_excel_without_auth(self, client):
        """Test that Excel upload requires authentication"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Description", "Unit"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        mime_type = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
        files = {"file": ("test.xlsx", excel_file, mime_type)}
        response = client.post("/api/parameters/upload-excel", files=files)
        assert response.status_code == status.HTTP_401_UNAUTHORIZED
