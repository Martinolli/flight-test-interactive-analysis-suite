"""
FTIAS Backend - Parameters Router
Test parameter management and Excel import endpoints
"""

import io
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import auth, schemas
from app.database import get_db
from app.models import TestParameter, User

router = APIRouter()


def _validate_min_max(min_value: Optional[float], max_value: Optional[float]) -> None:
    if min_value is not None and max_value is not None and min_value > max_value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="min_value cannot be greater than max_value",
        )


@router.post(
    "/",
    response_model=schemas.TestParameterResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_parameter(
    parameter: schemas.TestParameterCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _validate_min_max(parameter.min_value, parameter.max_value)

    existing = db.query(TestParameter).filter(TestParameter.name == parameter.name).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Parameter with name '{parameter.name}' already exists",
        )

    db_parameter = TestParameter(**parameter.model_dump())
    db.add(db_parameter)
    db.commit()
    db.refresh(db_parameter)
    return db_parameter


@router.get("/", response_model=List[schemas.TestParameterResponse])
async def get_parameters(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    system: Optional[str] = None,
    category: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    query = db.query(TestParameter)
    if search:
        query = query.filter(TestParameter.name.ilike(f"%{search}%"))
    if system:
        query = query.filter(TestParameter.system == system)
    if category:
        query = query.filter(TestParameter.category == category)

    return query.offset(skip).limit(limit).all()


@router.get("/{parameter_id:int}", response_model=schemas.TestParameterResponse)
async def get_parameter(
    parameter_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    parameter = db.query(TestParameter).filter(TestParameter.id == parameter_id).first()
    if not parameter:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Parameter not found"
        )
    return parameter


@router.put("/{parameter_id:int}", response_model=schemas.TestParameterResponse)
async def update_parameter(
    parameter_id: int,
    parameter: schemas.TestParameterUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    db_parameter = db.query(TestParameter).filter(TestParameter.id == parameter_id).first()
    if not db_parameter:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Parameter not found"
        )

    update_data = parameter.model_dump(exclude_unset=True)
    _validate_min_max(
        update_data.get("min_value", db_parameter.min_value),
        update_data.get("max_value", db_parameter.max_value),
    )

    if "name" in update_data:
        existing = (
            db.query(TestParameter)
            .filter(TestParameter.name == update_data["name"], TestParameter.id != parameter_id)
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Parameter with name '{update_data['name']}' already exists",
            )

    for key, value in update_data.items():
        setattr(db_parameter, key, value)

    db.add(db_parameter)
    db.commit()
    db.refresh(db_parameter)
    return db_parameter


@router.delete("/{parameter_id:int}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_parameter(
    parameter_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    parameter = db.query(TestParameter).filter(TestParameter.id == parameter_id).first()
    if not parameter:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Parameter not found"
        )

    db.delete(parameter)
    db.commit()
    return None


@router.post("/bulk", status_code=status.HTTP_201_CREATED)
async def bulk_create_parameters(
    request: schemas.BulkParametersCreateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    created = 0
    for param in request.parameters:
        _validate_min_max(param.min_value, param.max_value)

        exists = db.query(TestParameter).filter(TestParameter.name == param.name).first()
        if exists:
            continue
        db.add(TestParameter(**param.model_dump()))
        created += 1

    db.commit()
    return {"created": created}


@router.put("/bulk")
async def bulk_update_parameters(
    request: schemas.BulkParametersUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    updated = 0

    for item in request.parameters:
        db_parameter = db.query(TestParameter).filter(TestParameter.id == item.id).first()
        if not db_parameter:
            continue

        update_data = item.model_dump(exclude_unset=True)
        update_data.pop("id", None)

        _validate_min_max(
            update_data.get("min_value", db_parameter.min_value),
            update_data.get("max_value", db_parameter.max_value),
        )

        if "name" in update_data:
            existing = (
                db.query(TestParameter)
                .filter(TestParameter.name == update_data["name"], TestParameter.id != item.id)
                .first()
            )
            if existing:
                continue

        for key, value in update_data.items():
            setattr(db_parameter, key, value)
        db.add(db_parameter)
        updated += 1

    db.commit()
    return {"updated": updated}


@router.delete("/bulk", status_code=status.HTTP_204_NO_CONTENT)
async def bulk_delete_parameters(
    request: schemas.BulkParametersDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    if request.parameter_ids:
        db.query(TestParameter).filter(TestParameter.id.in_(request.parameter_ids)).delete(
            synchronize_session=False
        )
        db.commit()
    return None


@router.post("/upload-excel")
async def upload_parameters_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """
    Upload Excel file with test parameters.
    """
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be an Excel file",
        )

    try:
        contents = await file.read()
        workbook = load_workbook(filename=io.BytesIO(contents))
        sheet = workbook.active

        header_cells = [cell.value for cell in sheet[1]]
        headers = [
            (str(h).strip().lower() if h is not None else "") for h in header_cells
        ]

        def header_index(name: str) -> Optional[int]:
            name = name.strip().lower()
            try:
                return headers.index(name)
            except ValueError:
                return None

        name_idx = header_index("name")
        unit_idx = header_index("unit")
        if name_idx is None or unit_idx is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Missing required columns: Name and Unit",
            )

        # Optional columns
        desc_idx = header_index("description")
        system_idx = header_index("system")
        category_idx = header_index("category")
        min_idx = header_index("min value")
        max_idx = header_index("max value")

        created = 0
        updated = 0
        rows_processed = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[name_idx] is None or str(row[name_idx]).strip() == "":
                continue

            rows_processed += 1
            name = str(row[name_idx]).strip()
            unit = str(row[unit_idx]).strip() if row[unit_idx] is not None else ""
            description = (
                str(row[desc_idx]).strip() if desc_idx is not None and row[desc_idx] is not None else None
            )
            system = (
                str(row[system_idx]).strip() if system_idx is not None and row[system_idx] is not None else None
            )
            category = (
                str(row[category_idx]).strip() if category_idx is not None and row[category_idx] is not None else None
            )
            min_value = row[min_idx] if min_idx is not None else None
            max_value = row[max_idx] if max_idx is not None else None

            if isinstance(min_value, str) and min_value.strip() == "":
                min_value = None
            if isinstance(max_value, str) and max_value.strip() == "":
                max_value = None

            _validate_min_max(min_value, max_value)

            existing = db.query(TestParameter).filter(TestParameter.name == name).first()
            if existing:
                existing.description = description if description is not None else existing.description
                existing.unit = unit if unit is not None else existing.unit
                existing.system = system if system is not None else existing.system
                existing.category = category if category is not None else existing.category
                existing.min_value = min_value if min_value is not None else existing.min_value
                existing.max_value = max_value if max_value is not None else existing.max_value
                db.add(existing)
                updated += 1
                continue

            db.add(
                TestParameter(
                    name=name,
                    description=description,
                    unit=unit,
                    system=system,
                    category=category,
                    min_value=min_value,
                    max_value=max_value,
                )
            )
            created += 1

        db.commit()

        result = {
            "message": "Excel parameters uploaded successfully",
            "rows_processed": rows_processed,
            "parameters_created": created,
        }
        if updated:
            result["parameters_updated"] = updated
        return result

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing Excel file: {str(e)}",
        ) from e
